import openpyxl
from openpyxl.chart import BarChart, Reference
import os

class HistogramGenerator:
    def __init__(self, heterostructure):
        self.heterostructure = heterostructure

    def format_number(self, value):
        """Форматирует числа с использованием E в обычный вид."""
        if isinstance(value, (int, float)):
            return float(f"{value:.10f}")  # Убирает e и преобразует число
        return float(value)

    def extract_data_from_ai(self):
        """
        Извлекает данные из функции print_ai для записи в Excel.
        Создает отдельные списки для n, p и n+p.
        """
        alpha_i_data_n = []
        alpha_i_data_p = []
        alpha_i_data_np = []

        counts = 0
        for layer in self.heterostructure.HS:
            layer_name = layer[0]
            layer_thickness = layer[1]
            n_sum = 0.0
            p_sum = 0.0
            while counts < len(self.heterostructure.a_i) and self.heterostructure.a_i[counts][0] <= layer_thickness:
                n_sum += self.format_number(self.heterostructure.a_i[counts][1])
                p_sum += self.format_number(self.heterostructure.a_i[counts][2])
                counts += 1

            alpha_i_data_n.append([layer_name, n_sum])
            alpha_i_data_p.append([layer_name, p_sum])
            alpha_i_data_np.append([layer_name, n_sum + p_sum])
        return alpha_i_data_n, alpha_i_data_p, alpha_i_data_np

    def extract_data_from_rs(self):
        """
        Извлекает данные из функции print_rs для записи в Excel.
        Значения преобразуются из E-нотации в обычный вид.
        """
        rs_data = []
        total_resistance = sum(self.heterostructure.ro_i)
        rs_data.append(["Rs_total", self.format_number(total_resistance)])

        for i, layer in enumerate(self.heterostructure.HS):
            rs_data.append([layer[0], self.format_number(self.heterostructure.ro_i[i])])

        return rs_data

    def save_data_to_excel(self, data, sheet_name, output_excel):
        # Создание Excel файла
        if os.path.exists(output_excel):
            wb = openpyxl.load_workbook(output_excel)
        else:
            wb = openpyxl.Workbook()
            wb.remove(wb.active)

        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(title=sheet_name)

        # Запись данных в Excel
        ws.append(["Элемент", "Значение"])
        for row in data:
            ws.append(row)

        # Создание гистограммы
        chart = BarChart()
        chart.title = f"Гистограмма {sheet_name}"
        chart.x_axis.title = "Элемент"
        chart.y_axis.title = "Значение"

        data_range = Reference(ws, min_col=2, min_row=2, max_row=ws.max_row)
        labels = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
        chart.add_data(data_range, titles_from_data=False)
        chart.set_categories(labels)

        ws.add_chart(chart, "E5")

        # Сохранение Excel файла
        wb.save(output_excel)
        print(f"Данные и гистограмма сохранены в файл: {output_excel}")

class RsCalculatorWithGist:
    def __init__(self, sim_hs):
        self.sim_hs = sim_hs
        self.histogram_generator = HistogramGenerator(self.sim_hs)

    def generate_histograms(self):
        self.sim_hs.load_carriers_conc()
        self.sim_hs.load_near_field()
        self.sim_hs.calculate_ai()
        self.sim_hs.calculate_ros()

        # Генерация данных для AI гистограммы
        self.sim_hs.print_ros()
        ai_data_n, ai_data_p, ai_data_np = self.sim_hs.print_ai()
        rs_data = self.histogram_generator.extract_data_from_rs()

        # Сохранение данных и гистограмм в Excel
        output_file = "output.xlsx"
        self.histogram_generator.save_data_to_excel(rs_data, "RS", output_file)
        self.histogram_generator.save_data_to_excel(ai_data_n, "AI_n", output_file)
        self.histogram_generator.save_data_to_excel(ai_data_p, "AI_p", output_file)
        self.histogram_generator.save_data_to_excel(ai_data_np, "AI_n+p", output_file)
